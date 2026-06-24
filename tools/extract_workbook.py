#!/usr/bin/env python3
"""
extract_workbook.py  —  one-time/offline generator.

Reads the real Rosenbaum DCF workbook (Template + Completed) and emits a faithful,
JSON-ish JS data file (js/workbook-data.js) describing every cell of every sheet:
value, displayed string, formula, number format, classification (input / formula /
link / number / label / header), and merged ranges + column widths.

The platform's workbook renderer (js/workbook.js) reads this to recreate each sheet
exactly, with click-to-reveal formulas and colour-coded cell kinds.

Run:  python3 tools/extract_workbook.py
"""
import json, re, os, datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import DataTableFormula, ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)                       # IB Mastery Platform/
SRC  = os.path.dirname(ROOT)                        # finance/
TEMPLATE  = os.path.join(SRC, "DCF Analysis_3E_Template.xlsx")
COMPLETED = os.path.join(SRC, "DCF Analysis_3E_Completed.xlsx")
OUT = os.path.join(ROOT, "js", "workbook-data.js")

SHEETS = ["DCF", "AS1", "AS2", "NWC", "WACC"]

# Interior output ranges of two-variable data tables (sensitivity grids).
# These cells are blank in the Template and cached numbers in the Completed file,
# so they must NOT be treated as "inputs" — they are computed table outputs.
DATATABLE_RANGES = {
    "WACC": ["H27:L31"],
    "DCF":  ["D49:H53", "K49:O53", "D78:H82", "K78:O82",
             "D88:H92", "K88:O92", "G98:K102"],
}
SHEET_META = {
    "DCF":  ("ValueCo Corporation", "Discounted Cash Flow Analysis — output sheet"),
    "AS1":  ("ValueCo Corporation", "Assumptions Page 1 — Income Statement & Cash Flow"),
    "AS2":  ("ValueCo Corporation", "Assumptions Page 2 — Balance Sheet"),
    "NWC":  ("ValueCo Corporation", "Net Working Capital projections"),
    "WACC": ("ValueCo Corporation", "Weighted Average Cost of Capital analysis"),
}

def is_blank(v):
    return v is None or v == "" or v == " "

def expand_range(rng):
    """'H27:L31' -> set of cell addresses."""
    from openpyxl.utils import range_boundaries
    c1, r1, c2, r2 = range_boundaries(rng)
    out = set()
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            out.add(f"{get_column_letter(c)}{r}")
    return out

def datatable_cells(name):
    s = set()
    for rng in DATATABLE_RANGES.get(name, []):
        s |= expand_range(rng)
    return s

def formula_str(v):
    """Return a clean formula string or None. Data tables / arrays are flagged."""
    if isinstance(v, DataTableFormula):
        return "{=TABLE(...)}"          # two-variable data table
    if isinstance(v, ArrayFormula):
        return str(v.text)
    if isinstance(v, str) and v.startswith("="):
        return v
    return None

def references_other_sheet(formula, self_name):
    if not formula:
        return False
    # match  Name!  or  'Name'!  references
    for m in re.finditer(r"('([^']+)'|([A-Za-z_][A-Za-z0-9_ ]*))!", formula):
        name = m.group(2) or m.group(3)
        if name and name != self_name:
            return True
    return False

def count_decimals(nf):
    if "." in nf:
        frac = nf.split(".", 1)[1]
        # stop at first non 0/# char
        m = re.match(r"[0#]+", frac)
        return len(m.group(0)) if m else 0
    return 0

def fmt_display(v, nf):
    """Pragmatic Excel-like display string."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.strftime("%m/%d/%Y").lstrip("0").replace("/0", "/")
    nf = nf or "General"
    low = nf.lower()
    try:
        if "%" in nf:
            d = count_decimals(nf)
            return f"{v*100:.{d}f}%"
        if '"x"' in low or low.endswith("x") or re.search(r'0\.?0*"?x', low):
            d = count_decimals(nf) or 1
            return f"{v:.{d}f}x"
        if "$" in nf or "#,##0" in nf or "," in nf:
            d = count_decimals(nf)
            s = f"{abs(v):,.{d}f}"
            if "$" in nf:
                s = "$" + s
            return f"({s})" if v < 0 else s
        # general
        if isinstance(v, float):
            if v.is_integer():
                return f"{int(v):,}"
            return f"{v:,.2f}"
        return f"{v:,}" if isinstance(v, int) else str(v)
    except Exception:
        return str(v)

def used_bounds(ws_f):
    """True used range (openpyxl max_* can be inflated)."""
    max_r = 0; max_c = 0
    for row in ws_f.iter_rows():
        for c in row:
            if not is_blank(c.value):
                if c.row > max_r: max_r = c.row
                if c.column > max_c: max_c = c.column
    return max_r, min(max_c, 30)        # cap at col 30 (AD) — Output runs to AB

def classify(addr, val_completed, formula, is_input, self_name):
    if is_input:
        return "input"
    if formula:
        if references_other_sheet(formula, self_name):
            return "link"
        if formula == "{=TABLE(...)}":
            return "datatable"
        return "formula"
    if isinstance(val_completed, str):
        return "label"
    if val_completed is not None:
        return "number"             # hardcoded constant that is the same in template (given)
    return "empty"

def extract_sheet(out_key, tname, cname, wb_tmpl_f, wb_comp_f, wb_comp_v, dt_ranges, meta):
    ws_tf = wb_tmpl_f[tname]
    ws_cf = wb_comp_f[cname]
    ws_cv = wb_comp_v[cname]
    max_r, max_c = used_bounds(ws_cf)
    # union with template bounds (template may have rows completed lacks, e.g. extra)
    tr, tc = used_bounds(ws_tf)
    max_r = max(max_r, tr); max_c = max(max_c, tc)
    dt_cells = set()
    for rng in dt_ranges:
        dt_cells |= expand_range(rng)

    cells = {}
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            addr = f"{get_column_letter(c)}{r}"
            comp_raw = ws_cf.cell(row=r, column=c).value           # formula or constant
            comp_val = ws_cv.cell(row=r, column=c).value           # computed value
            tmpl_raw = ws_tf.cell(row=r, column=c).value
            nf = ws_cf.cell(row=r, column=c).number_format

            formula = formula_str(comp_raw)
            # display value: prefer computed value; for labels comp_val == text
            disp = fmt_display(comp_val if comp_val is not None else
                               (comp_raw if not formula else None), nf)

            # input detection: blank/zero in template, but filled (non-formula constant) in completed
            comp_is_const = (formula is None) and (not is_blank(comp_raw)) and not isinstance(comp_raw, str)
            tmpl_blank = is_blank(tmpl_raw) or tmpl_raw == 0
            is_input = comp_is_const and tmpl_blank

            if is_blank(comp_raw) and is_blank(comp_val) and is_blank(tmpl_raw):
                continue

            kind = classify(addr, comp_val, formula, is_input, cname)
            if addr in dt_cells and kind in ("input", "number", "formula", "datatable"):
                kind = "datatable"          # sensitivity-table output, not a user input
            rec = {"v": disp, "k": kind}
            if formula:
                rec["f"] = formula
            cell_obj = ws_cf.cell(row=r, column=c)
            if cell_obj.font and cell_obj.font.bold:
                rec["b"] = 1
            # alignment hint: explicit right/center else default
            al = cell_obj.alignment.horizontal if cell_obj.alignment else None
            if al in ("right", "center", "left"):
                rec["a"] = al[0]            # 'r' | 'c' | 'l'
            cells[addr] = rec

    merges = [str(m) for m in ws_cf.merged_cells.ranges]
    widths = {}
    for col_letter, dim in ws_cf.column_dimensions.items():
        if dim.width:
            widths[col_letter] = round(dim.width, 1)

    company, subtitle = meta
    return {
        "title": company,
        "subtitle": subtitle,
        "maxRow": max_r,
        "maxCol": max_c,
        "cells": cells,
        "merges": merges,
        "widths": widths,
    }

def build(template_path, completed_path, sheets, dt_map, out_path, label):
    """sheets: list of (out_key, tmpl_sheet, comp_sheet, company, subtitle)."""
    wb_tf = openpyxl.load_workbook(template_path, data_only=False)
    wb_cf = openpyxl.load_workbook(completed_path, data_only=False)
    wb_cv = openpyxl.load_workbook(completed_path, data_only=True)
    out = {}
    print(f"== {label} ==")
    for (key, tname, cname, company, sub) in sheets:
        out[key] = extract_sheet(key, tname, cname, wb_tf, wb_cf, wb_cv,
                                 dt_map.get(key, []), (company, sub))
        nin = sum(1 for c in out[key]["cells"].values() if c["k"] == "input")
        print(f"  {key}: {len(out[key]['cells'])} cells, {nin} inputs, "
              f"{out[key]['maxRow']}x{out[key]['maxCol']}")
    js = ("/* AUTO-GENERATED by tools/extract_workbook.py — do not edit by hand. */\n"
          "const WORKBOOK = " + json.dumps(out, ensure_ascii=False, indent=1) + ";\n"
          "if (typeof module !== 'undefined') module.exports = { WORKBOOK };\n")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(js)
    print(f"  -> wrote {out_path}  ({os.path.getsize(out_path):,} bytes)\n")

def main():
    # ---- DCF workbook (5 tabs, same template/completed sheet names) ----
    dcf_sheets = [(n, n, n) + SHEET_META[n] for n in SHEETS]
    build(TEMPLATE, COMPLETED, dcf_sheets, DATATABLE_RANGES, OUT, "DCF workbook")

    # ---- Comparable Companies workbook ----
    cdir = os.path.join(ROOT, "js", "comps-data.js")
    comps_tmpl = os.path.join(SRC, "Comparable Companies_3E_Template.xlsx")
    comps_comp = os.path.join(SRC, "Comparable Companies_3E_Completed.xlsx")
    # (out_key, template_sheet, completed_sheet, company, subtitle)
    # The company "Spread" uses the blank template sheet CompCo 1 vs the filled JDG
    # so the positional diff still flags the analyst's hardcoded inputs.
    comps_sheets = [
        ("List",   "List",          "List",          "ValueCo Corporation", "List of Comparable Companies — the universe"),
        ("Spread", "CompCo 1",      "JDG",           "Gasparro Corp. (JDG)", "Company Input Page — spreading one comparable"),
        ("Bench",  "Benchmarking 2", "Benchmarking 2", "ValueCo Corporation", "Benchmarking — returns, leverage & coverage ratios"),
        ("Output", "Ouput",         "Ouput",         "ValueCo Corporation", "Comparable Companies Analysis — output summary"),
    ]
    build(comps_tmpl, comps_comp, comps_sheets, {}, cdir, "Comparable Companies workbook")

if __name__ == "__main__":
    main()
